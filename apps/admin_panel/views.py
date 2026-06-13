
#/ Замена / Alternative to Django admin
#? Роли / Roles: superuser = admin, expansion_manager = stores, staff = products/orders
#? Roles: superuser, expansion_manager, staff

import csv

from functools import wraps
from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import Count, Sum, Q
from django.contrib.auth import get_user_model
from django.conf import settings
from django.contrib import messages
from config import logger
from django.utils.text import slugify

from apps.catalog.models import Product, Category, ExchangeRate, CategorySpecField
from apps.catalog.forms import ProductForm, CategoryForm
from apps.orders.models import Order, OrderItem
from apps.reviews.models import Review
from apps.stores.models import Store, StoreStaff, StoreProduct
from apps.stores.forms import StoreForm
from apps.orders.tasks import send_order_status_update

User = get_user_model()



def admin_required(view_func):
    #* Декоратор / Decorator: checks is_staff + 2FA, no email check
    #* Если 2FA не настроена / If 2FA is not set up — redirect to setup
    @wraps(view_func)
    def _wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(
                f'{reverse(settings.LOGIN_URL)}?next={request.path}'
            )
        if (request.user.is_staff or request.user.is_expansion_manager) == False:
            return redirect('accounts:login')
        #! Без 2FA / No 2FA — force setup (google authenticator)
        if not request.user.totp_secret:
            return redirect('accounts:otp_setup')
        return view_func(request, *args, **kwargs)
    return _wrapper


def expansion_manager_required(view_func):
    #* Декоратор / Decorator for store management (admin + expansion_manager)
    @wraps(view_func)
    def _wrapper(request, *args, **kwargs):
        if not (request.user.is_superuser or request.user.is_expansion_manager):
            messages.error(request, "Insufficient permissions.")
            return redirect('admin_panel:dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapper


# ─── Дашборд / Dashboard ───

@admin_required
def dashboard(request):
    total_products = Product.objects.count()
    total_orders = Order.objects.count()
    total_users = User.objects.count()
    total_stores = Store.objects.count()
    total_revenue = Order.objects.filter(paid = True).aggregate(
        Sum('total_cost')
    )['total_cost__sum']  or 0

    new_orders = Order.objects.filter(status='new')[:10]
    low_stock = Product.objects.filter(stock__lt=5, is_available=True)[:10]

    logger.debug(
        f'Dashboard: {total_products} products, {total_orders} orders, '
        f'£{total_revenue} revenue'
    )

    return render(request, 'admin_panel/dashboard.html', {
        'total_products': total_products,
        'total_orders': total_orders,
        'total_users': total_users,
        'total_stores': total_stores,
        'total_revenue': total_revenue,
        'new_orders': new_orders,
        'low_stock': low_stock,
    })


# ─── Товары / Products ───

@admin_required
def product_list(request):
    products = Product.objects.select_related('category').all()
    return render(
        request, 'admin_panel/product_list.html', {'products': products}
        )


@admin_required
def product_create(request):
    category_id = request.GET.get('category')
    if not category_id and request.method != 'POST':
        # Step 1: choose a category / Шаг 1: выбор категории
        categories = Category.objects.filter(is_active=True)
        return render(request, 'admin_panel/product_choose_category.html', {
            'categories': categories,
        })
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            logger.info('Product created: id={}, name={}, price={} by {}',
                         product.id, product.name, product.price, request.user.username)
            messages.success(request, "Product created.")
            return redirect('admin_panel:product_list')
    else:
        initial = {}
        try:
            initial['category'] = int(category_id)
        except (ValueError, TypeError):
            pass
        form = ProductForm(initial=initial)
    return render(
        request, 'admin_panel/product_form.html',
        {'form': form, 'title': 'New product'}
    )


@admin_required
def product_edit(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            logger.info('Product updated: id={}, name={} by {}',
                         product.id, product.name, request.user.username)
            messages.success(request, "Product updated.")
            return redirect('admin_panel:product_list')
    else:
        form = ProductForm(instance=product)
    return render(
        request, 'admin_panel/product_form.html',
        {'form': form, 'title': 'Edit product'}
    )


@admin_required
def product_delete(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        logger.info('Product deleted: id={}, name={} by {}',
                     product.id, product.name, request.user.username)
        product.delete()
        messages.success(request, 'Product deleted.')
        return redirect('admin_panel:product_list')
    return render(
        request, 'admin_panel/product_confirm_delete.html',
        {'object': product, 'title': 'Delete product'}
    )


@admin_required
def product_export_csv(request):
    #* Экспорт / Export products with category
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    response.write('\ufeff'.encode('utf-8'))

    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Name', 'Category', 'Price', 'Old price',
        'Discount %', 'Stock', 'Available', 'Created',
    ])

    products = Product.objects.select_related('category').all()
    for p in products:
        writer.writerow([
            p.id, p.name, p.category.name if p.category else '',
            p.price, p.old_price or '', p.discount_percent,
            p.stock, 'Yes' if p.is_available else 'No', p.created,
        ])

    return response


# ─── Категории / Categories ───

@admin_required
def category_list(request):
    categories = Category.objects.all()
    return render(
        request, 'admin_panel/category_list.html',
        {'categories': categories}
    )


@admin_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category created.')
            return redirect('admin_panel:category_list')
    else:
        form = CategoryForm()
    return render(
        request, 'admin_panel/category_form.html',
        {'form': form, 'title': 'New category'}
    )


@admin_required
def category_edit(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated.')
            return redirect('admin_panel:category_list')
    else:
        form = CategoryForm(instance=category)
    return render(
        request, 'admin_panel/category_form.html',
        {'form': form, 'title': 'Edit category'}
    )


@admin_required
def category_delete(request, category_id):
    #* Удаление / Delete category (POST only)
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted.')
    return redirect('admin_panel:category_list')


# ─── Category Spec Fields / Поля характеристик категорий ───

@admin_required
def category_spec_field_list(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    fields = CategorySpecField.objects.filter(category=category)
    return render(request, 'admin_panel/category_spec_field_list.html', {
        'category': category, 'fields': fields,
    })


@admin_required
def category_spec_field_edit(request, category_id, field_id=None):
    category = get_object_or_404(Category, id=category_id)
    field_obj = get_object_or_404(CategorySpecField, id=field_id, category=category) if field_id else None
    if request.method == 'POST':
        from apps.catalog.forms import CategorySpecFieldForm
        form = CategorySpecFieldForm(request.POST, instance=field_obj)
        if form.is_valid():
            f = form.save(commit=False)
            f.category = category
            f.save()
            messages.success(request, 'Spec field saved.')
            return redirect('admin_panel:category_spec_field_list', category_id=category.id)
    else:
        from apps.catalog.forms import CategorySpecFieldForm
        form = CategorySpecFieldForm(instance=field_obj)
    return render(request, 'admin_panel/category_spec_field_form.html', {
        'form': form, 'category': category, 'field_obj': field_obj,
    })


@admin_required
def category_spec_field_delete(request, category_id, field_id):
    field_obj = get_object_or_404(CategorySpecField, id=field_id, category_id=category_id)
    if request.method == 'POST':
        field_obj.delete()
        messages.success(request, 'Spec field deleted.')
    return redirect('admin_panel:category_spec_field_list', category_id=category_id)


# ─── Заказы / Orders ───

@admin_required
def order_list(request):
    orders = Order.objects.select_related('user', 'store').all()
    return render(
        request, 'admin_panel/order_list.html', {'orders': orders}
    )


@admin_required
def order_detail(request, order_id):
    order = get_object_or_404(
        Order.objects.select_related('user', 'store'), id=order_id
    )
    items = order.items.all()
    return render(
        request, 'admin_panel/order_detail.html',
        {'order': order, 'items': items}
    )


@admin_required
def order_update_status(request, order_id):
    #* Меняет статус / Change order status and send notification
    order = get_object_or_404(Order, id=order_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save(update_fields=['status'])
            messages.success(request, f'Order #{order_id} status updated.')
        else:
            messages.error(request, "Invalid status.")
    return redirect('admin_panel:order_detail', order_id=order_id)


# ─── Пользователи / Users ───

@admin_required
def user_list(request):
    users = User.objects.annotate(
        orders_count=Count('orders')
    ).all()
    return render(
        request, 'admin_panel/user_list.html', {'users': users}
    )


# ─── Отзывы / Reviews ───

@admin_required
def review_list(request):
    reviews = Review.objects.select_related('user', 'product').all()
    return render(
        request, 'admin_panel/review_list.html', {'reviews': reviews}
    )


@admin_required
def review_moderate(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    review.is_moderated = True
    review.save(update_fields=['is_moderated'])
    messages.success(request, 'Review published.')
    return redirect('admin_panel:review_list')


@admin_required
def review_delete(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    if request.method == 'POST':
        review.delete()
        messages.success(request, 'Review deleted.')
        return redirect('admin_panel:review_list')
    return render(
        request, 'admin_panel/review_confirm_delete.html',
        {'object': review}
    )


# ─── Магазины / Stores ───

@expansion_manager_required
def store_list(request):
    stores = Store.objects.annotate(
        staff_count=Count('staff', filter=Q(staff__is_active=True)),
        product_count=Count('products'),
    ).all()
    return render(
        request, 'admin_panel/stores/store_list.html', {'stores': stores}
    )


@expansion_manager_required
def store_create(request):
    if request.method == 'POST':
        form = StoreForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Store created.')
            return redirect('admin_panel:store_list')
    else:
        form = StoreForm()
    return render(
        request, 'admin_panel/stores/store_form.html',
        {'form': form, 'title': 'New store'}
    )


@expansion_manager_required
def store_edit(request, store_id):
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            messages.success(request, 'Store updated.')
            return redirect('admin_panel:store_list')
    else:
        form = StoreForm(instance=store)
    return render(
        request, 'admin_panel/stores/store_form.html',
        {'form': form, 'title': 'Edit store'}
    )


@expansion_manager_required
def store_staff(request, store_id):
    #* Управление / Manage store staff
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        role = request.POST.get('role', 'seller')
        if user_id:
            # add or update
            StoreStaff.objects.update_or_create(
                store=store, user_id=user_id,
                defaults={'role': role,'is_active': True},
            )
            messages.success(request, 'Staff assigned.')
        return redirect('admin_panel:store_staff', store_id=store.id)

    staff = StoreStaff.objects.filter(store=store).select_related('user')
    available_users = User.objects.filter(is_staff=True)
    return render(request, 'admin_panel/stores/store_staff.html', {
        'store': store, 'staff': staff, 'available_users': available_users,
    })


@expansion_manager_required
def store_stock(request, store_id):
    #* Управление / Manage store stock
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        stock = request.POST.get('stock', 0)
        if product_id:
            StoreProduct.objects.update_or_create(
                store=store, product_id=product_id,
                defaults={'stock': stock},
            )
            messages.success(request, 'Stock updated.')
        return redirect('admin_panel:store_stock', store_id=store.id)

    stocks = StoreProduct.objects.filter(store=store).select_related('product')
    all_products = Product.objects.filter(is_available=True)
    return render(request, 'admin_panel/stores/store_stock.html', {
        'store': store, 'stocks': stocks, 'all_products': all_products,
    })


@expansion_manager_required
def store_delete(request, store_id):
    #* Удаление / Delete store (POST only)
    store = get_object_or_404(Store, id=store_id)
    if request.method == 'POST':
        store.delete()
        messages.success(request, 'Store deleted.')
    return redirect('admin_panel:store_list')


# ─── Exchange Rates / Курсы валют ───

@admin_required
def exchange_rate_list(request):
    rates = ExchangeRate.objects.all()
    return render(request, 'admin_panel/exchange_rate_list.html', {'rates': rates})


@admin_required
def exchange_rate_edit(request, rate_id=None):
    rate = get_object_or_404(ExchangeRate, id=rate_id) if rate_id else None
    if request.method == 'POST':
        from apps.catalog.forms import ExchangeRateForm
        form = ExchangeRateForm(request.POST, instance=rate)
        if form.is_valid():
            form.save()
            messages.success(request, 'Exchange rate saved.')
            return redirect('admin_panel:exchange_rate_list')
    else:
        from apps.catalog.forms import ExchangeRateForm
        form = ExchangeRateForm(instance=rate)
    return render(request, 'admin_panel/exchange_rate_form.html', {
        'form': form, 'rate': rate,
    })


# ─── Аналитика / Analytics ───

@admin_required
def analytics_dashboard(request):
    #* Статистика / Sales stats: total revenue, order count, top products
    total_revenue = Order.objects.filter(
        ~Q(status='cancelled')
    ).aggregate(Sum('total_cost'))['total_cost__sum'] or 0

    total_orders = Order.objects.count()
    avg_order_value = total_revenue/total_orders if total_orders else 0

    # Топ товаров по продажам / Top products by sales
    top_products = OrderItem.objects.values(
        'product_name'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('product_price') * Sum('quantity'),
    ).order_by('-total_sold')[:10]

    return render(request, 'admin_panel/analytics/dashboard.html', {
        'total_revenue': total_revenue,
        'total_orders': total_orders,
        'avg_order_value': round(avg_order_value, 2),
        'top_products': top_products,
    })


def _build_excel_report(orders_qs, period_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    # styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    money_fmt = '#,##0.00'

    # title row
    ws.merge_cells('A1:H1')
    ws['A1'] = f'MegaShop — {period_label}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # period sub-title
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    # headers
    headers = ['Order #', 'Date', 'Customer', 'Phone', 'Delivery', 'Items', 'Status', 'Total (£)']
    row_num = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # data
    for order in orders_qs.select_related('pickup_store').prefetch_related('items'):
        row_num += 1
        items_str = ', '.join([f'{i.product_name} x{i.quantity}' for i in order.items.all()])
        delivery_str = order.pickup_store.name if order.delivery_type == 'pickup' and order.pickup_store else f'{order.city}, {order.address}'
        ws.cell(row=row_num, column=1, value=order.id).border = thin_border
        ws.cell(row=row_num, column=2, value=order.created.strftime('%d.%m.%Y')).border = thin_border
        ws.cell(row=row_num, column=3, value=f'{order.first_name} {order.last_name}').border = thin_border
        ws.cell(row=row_num, column=4, value=order.phone).border = thin_border
        ws.cell(row=row_num, column=5, value=delivery_str).border = thin_border
        ws.cell(row=row_num, column=6, value=items_str).border = thin_border
        ws.cell(row=row_num, column=7, value=order.get_status_display()).border = thin_border
        cell = ws.cell(row=row_num, column=8, value=float(order.total_cost))
        cell.number_format = money_fmt
        cell.border = thin_border

    # summary row
    row_num += 2
    total_rev = orders_qs.aggregate(s=Sum('total_cost'))['s'] or 0
    total_orders_count = orders_qs.count()
    ws.cell(row=row_num, column=1, value='TOTAL').font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=2, value=total_orders_count).font = Font(bold=True)
    ws.cell(row=row_num, column=8, value=float(total_rev)).font = Font(bold=True, size=11)
    ws.cell(row=row_num, column=8).number_format = money_fmt

    # column widths
    widths = [10, 14, 22, 16, 30, 40, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb


@admin_required
def export_sales_excel(request, period):
    from dateutil.relativedelta import relativedelta
    from django.utils import timezone
    now = timezone.now()
    if period == 'month':
        start_date = timezone.make_aware(datetime(now.year, now.month, 1))
        end_date = start_date + relativedelta(months=1)
        label = f'Sales Report — {now.strftime("%B %Y")}'
        filename = f'megashop_sales_{now.strftime("%Y_%m")}.xlsx'
    elif period == 'year':
        start_date = timezone.make_aware(datetime(now.year, 1, 1))
        end_date = timezone.make_aware(datetime(now.year + 1, 1, 1))
        label = f'Sales Report — {now.year}'
        filename = f'megashop_sales_{now.year}.xlsx'
    else:
        messages.error(request, 'Invalid period.')
        return redirect('admin_panel:analytics_dashboard')

    orders = Order.objects.filter(
        created__gte=start_date,
        created__lt=end_date,
    ).order_by('-created')

    wb = _build_excel_report(orders, label)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
